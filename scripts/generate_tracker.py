#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a reusable job-application tracker (.xlsx) from a priority config.

Usage:
    python scripts/generate_tracker.py \
        --config priority_config.json \
        --output 求职投递台账.xlsx

Dependency: openpyxl  (pip install openpyxl)
"""
import argparse
import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "priority_config.example.json")

HEADERS = [
    "编号", "投递日期", "投递截止时间", "公司名称", "岗位名称",
    "投递渠道", "JD/投递链接", "是否官网自建账号", "账号", "密码", "所用简历版本",
    "确认状态", "投递状态", "备注",
]

WIDTHS = [6, 12, 20, 20, 12, 12, 32, 14, 14, 18, 20, 10, 10, 24]


def build_tracker(cfg: dict, out_path: str) -> None:
    wb = openpyxl.Workbook()

    # ---------- Sheet 1: 投递记录 ----------
    ws = wb.active
    ws.title = "投递记录"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # 下拉校验（可选便利项）
    dv_confirm = DataValidation(type="list", formula1='"待确认,已确认,已提交,已回绝"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"已投,面试中,Offer,已拒,无回复"', allow_blank=True)
    dv_acc = DataValidation(type="list", formula1='"是,否,待定"', allow_blank=True)
    ws.add_data_validation(dv_confirm)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_acc)
    dv_confirm.add("K2:K1000")
    dv_status.add("L2:L1000")
    dv_acc.add("G2:G1000")

    # ---------- Sheet 2: 优先级配置参考（由配置生成） ----------
    ws2 = wb.create_sheet("优先级配置参考")

    industry_max = max(item["level"] for item in cfg["industry_priority"])
    role_max = max(item["level"] for item in cfg["role_priority"])
    w_i = cfg.get("weights", {}).get("industry", 10)
    w_r = cfg.get("weights", {}).get("role", 5)

    def score_formula(kind: str, level: int) -> str:
        base = (industry_max + 1 - level) * w_i if kind == "industry" else (role_max + 1 - level) * w_r
        return f"{base}"

    rows = [
        ["【行业优先级阶梯】（level 越小越优先）"],
        ["level", "行业", "匹配关键词", "行业得分"],
    ]
    for item in sorted(cfg["industry_priority"], key=lambda x: x["level"]):
        rows.append([item["level"], item["name"], " / ".join(item.get("match_keywords", [])), score_formula("industry", item["level"])])

    rows.append([])
    rows.append(["【岗位优先级阶梯】（level 越小越优先）"])
    rows.append(["level", "岗位", "匹配关键词", "岗位得分"])
    for item in sorted(cfg["role_priority"], key=lambda x: x["level"]):
        rows.append([item["level"], item["name"], " / ".join(item.get("match_keywords", [])), score_formula("role", item["level"])])

    rows.append([])
    rows.append(["【综合优先级分】"])
    rows.append([f"综合分 = 行业得分 + 岗位得分；行业得分 = ({industry_max}+1-level)×{w_i}，岗位得分 = ({role_max}+1-level)×{w_r}。分越高越优先。"])
    rows.append([f"每日流水线：搜岗 → 按配置归类打分 → 取 Top {cfg.get('max_batch', 20)} 写入「投递记录」(确认状态=待确认) → 人工确认/提交 → 回写状态。"])
    rows.append(["安全提示：密码列为明文，请勿公开或外发；本台账仅作个人求职记录。"])
    rows.append(["修改优先级：编辑 priority_config.json 后重新生成本表 / 通知自动化重新读取。"])

    for row in rows:
        ws2.append(row)

    for r in range(1, ws2.max_row + 1):
        val = ws2.cell(row=r, column=1).value
        if val and (str(val).startswith("【") or str(val) in ("level", "行业", "岗位")):
            ws2.cell(row=r, column=1).font = Font(bold=True, color="1F4E78")

    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 12

    parent = os.path.dirname(os.path.abspath(out_path))
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(out_path)
    print(f"saved: {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description="从优先级配置生成求职投递台账")
    ap.add_argument("--config", default=DEFAULT_CONFIG, help="优先级配置 JSON 路径")
    ap.add_argument("--output", default="求职投递台账.xlsx", help="输出 xlsx 路径")
    args = ap.parse_args()

    if not os.path.exists(args.config):
        print(f"[error] config not found: {args.config}", file=sys.stderr)
        sys.exit(1)

    with open(args.config, encoding="utf-8") as f:
        cfg = json.load(f)

    build_tracker(cfg, args.output)


if __name__ == "__main__":
    main()
